"""
FeedGen v4.0 - AI Feed Text Generator
======================================
Universal web service for generating optimized product titles, descriptions,
and missing attributes for Google Merchant Center feeds using Claude AI.

v4.0 (UX flow update, per CLAUDE_CODE_CONTEXT spec):
- Preview step: generate a small sample (default 5) BEFORE the full run
- Confirmation step with real cost estimate (based on actual preview token usage)
- Cancellation: stop a running job, partial results are saved and downloadable
- Structured per-product errors/warnings in job results
(settings persistence and no-reload restart live on the frontend)

v3.1 base (kept): token auth + tenant isolation, AsyncAnthropic parallel
generation, strict input validation, size/product limits, defusedxml,
attribute-aware deduplication, prompt caching with warm-up request.

Usage:
    pip install -r requirements.txt
    export ANTHROPIC_API_KEY=sk-ant-...
    export ACCESS_TOKENS=token1:client1,token2:client2   # optional, omit to allow all
    # optional tuning:
    export MAX_FEED_SIZE_MB=50
    export MAX_UNIQUE_PRODUCTS=3000
    export GENERATION_CONCURRENCY=5
    export COOKIE_INSECURE=1   # only for local http development
    uvicorn app:app --host 0.0.0.0 --port 8000
"""

import os
import csv
import json
import re
import time
import uuid
import asyncio
import logging
import tempfile
import secrets
import traceback
import xml.etree.ElementTree as ET  # writing only (safe)
from pathlib import Path
from datetime import datetime
from collections import Counter

import defusedxml.ElementTree as DET  # parsing untrusted XML

from fastapi import FastAPI, UploadFile, File, Form, HTTPException, BackgroundTasks, Request
from fastapi.responses import JSONResponse, RedirectResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.middleware.base import BaseHTTPMiddleware

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s [feedgen] %(message)s")
logger = logging.getLogger("feedgen")

app = FastAPI(title="FeedGen", version="4.0")

BASE_DIR = Path(__file__).parent
STATIC_DIR = BASE_DIR / "static"
TEMPLATES_DIR = BASE_DIR / "templates"
STATIC_DIR.mkdir(exist_ok=True)
TEMPLATES_DIR.mkdir(exist_ok=True)

app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")
templates = Jinja2Templates(directory=str(TEMPLATES_DIR))

UPLOAD_DIR = Path(tempfile.gettempdir()) / "feedgen"
UPLOAD_DIR.mkdir(exist_ok=True)

# =============================================================================
# CONFIG / LIMITS
# =============================================================================

MAX_FEED_SIZE = int(os.environ.get("MAX_FEED_SIZE_MB", "50")) * 1024 * 1024
MAX_CONFIG_SIZE = 1 * 1024 * 1024
MAX_UNIQUE_PRODUCTS = int(os.environ.get("MAX_UNIQUE_PRODUCTS", "3000"))
GENERATION_CONCURRENCY = max(1, int(os.environ.get("GENERATION_CONCURRENCY", "5")))
FILE_TTL_HOURS = 24
PREVIEW_DEFAULT_SIZE = 5
PREVIEW_MAX_SIZE = 10

ALLOWED_INPUT_EXTS = {"xlsx", "csv", "xml"}
ALLOWED_OUTPUT_FORMATS = {"xlsx", "csv", "xml"}
ALLOWED_LANGUAGES = {"uk", "en"}
ALLOWED_MODELS = {
    "claude-sonnet-5",
    "claude-sonnet-4-6",
    "claude-haiku-4-5",
    "claude-haiku-4-5-20251001",
}
CACHE_SUPPORTED_MODELS = set(ALLOWED_MODELS)

# Models that reject non-default temperature/top_p/top_k with a 400 error.
# Claude Sonnet 5 introduced this behavior change vs Sonnet 4.6 - it uses
# adaptive thinking by default and no longer accepts manual sampling tuning.
SONNET5_NO_SAMPLING_PARAMS = {"claude-sonnet-5"}

# Approx public pricing, USD per 1M tokens (input, output). Used ONLY for a
# rough pre-run estimate shown to the user; actual billing is on Anthropic side.
# Sonnet 5 intro pricing is $2/$10 through Aug 31 2026, then $3/$15 standard.
# Using the standard (post-intro) rate here so the estimate doesn't quietly
# become too low once the intro window ends.
MODEL_PRICING = {
    "claude-sonnet-5": (3.0, 15.0),
    "claude-sonnet-4-6": (3.0, 15.0),
    "claude-haiku-4-5": (1.0, 5.0),
    "claude-haiku-4-5-20251001": (1.0, 5.0),
}

FILE_ID_RE = re.compile(r"^[a-f0-9]{32}$")
JOB_ID_RE = re.compile(r"^[a-f0-9]{12}$")

# In-memory state (ephemeral by design; survives within one deploy)
jobs = {}            # job_id -> {...}
uploaded_files = {}  # file_id -> {"path", "ext", "client", "filename", "created"}


_last_cleanup_ts = 0.0
CLEANUP_MIN_INTERVAL = 300  # не частіше раз на 5 хвилин, і тільки в фоновому треді


def _cleanup_old_files_sync():
    """Синхронна частина чистки - виконується в окремому треді, не в event loop."""
    cutoff = time.time() - FILE_TTL_HOURS * 3600
    removed = 0
    try:
        for f in UPLOAD_DIR.iterdir():
            try:
                if f.is_file() and f.stat().st_mtime < cutoff:
                    f.unlink(missing_ok=True)
                    removed += 1
            except OSError:
                pass
    except OSError:
        pass
    stale = [fid for fid, meta in uploaded_files.items() if not Path(meta["path"]).exists()]
    for fid in stale:
        uploaded_files.pop(fid, None)
    if removed:
        logger.info("cleanup: removed %d old temp files", removed)


async def cleanup_old_files():
    """Чистка старих файлів. Раніше виконувалась синхронно на КОЖЕН /api/analyze
    і блокувала event loop, через що інші запити (навіть аналіз крихітних файлів)
    зависали в черзі за нею. Тепер: (1) не частіше раз на 5 хв, (2) винесена
    в окремий тред через asyncio.to_thread, щоб не тримати loop."""
    global _last_cleanup_ts
    now = time.time()
    if now - _last_cleanup_ts < CLEANUP_MIN_INTERVAL:
        return
    _last_cleanup_ts = now
    await asyncio.to_thread(_cleanup_old_files_sync)


# =============================================================================
# TOKEN AUTH
# =============================================================================

def load_tokens():
    raw = os.environ.get("ACCESS_TOKENS", "").strip()
    if not raw:
        return {}
    tokens = {}
    for part in raw.split(","):
        part = part.strip()
        if ":" in part:
            token, label = part.split(":", 1)
            if token.strip():
                tokens[token.strip()] = label.strip()
        elif part:
            tokens[part] = part[:8] + "..."
    return tokens


TOKENS = load_tokens()
AUTH_ENABLED = bool(TOKENS)
COOKIE_SECURE = os.environ.get("COOKIE_INSECURE", "") != "1"


def verify_token(token) -> str | None:
    """Returns client label if token is valid, None otherwise. Constant-time compare."""
    if not AUTH_ENABLED:
        return "open"
    if not token or not isinstance(token, str):
        return None
    label = None
    for t, lbl in TOKENS.items():
        if secrets.compare_digest(token, t):
            label = lbl
    return label


def get_client(request: Request) -> str | None:
    token = request.headers.get("X-Access-Token") or request.cookies.get("feedgen_token")
    return verify_token(token)


PUBLIC_PATHS = {"/api/health", "/login", "/api/login"}


class AuthMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        path = request.url.path
        if path in PUBLIC_PATHS or path.startswith("/static/"):
            return await call_next(request)
        if not AUTH_ENABLED:
            return await call_next(request)
        if not get_client(request):
            if path.startswith("/api/"):
                return JSONResponse({"detail": "Unauthorized"}, status_code=401)
            return RedirectResponse(url="/login")
        return await call_next(request)


app.add_middleware(AuthMiddleware)

LOGIN_WINDOW = 300
LOGIN_MAX_ATTEMPTS = 10
login_attempts = {}


# =============================================================================
# GOOGLE MERCHANT CENTER ATTRIBUTE SCHEMA
# =============================================================================

GMC_ATTRIBUTES = {
    "id": {"aliases": ["id", "item_id", "offer_id", "товар_id"], "generatable": False, "label": "ID"},
    "title": {"aliases": ["title", "name", "назва", "найменування"], "generatable": False, "label": "Title"},
    "description": {"aliases": ["description", "desc", "опис"], "generatable": False, "label": "Description"},
    "link": {"aliases": ["link", "url", "посилання"], "generatable": False, "label": "Link"},
    "image_link": {"aliases": ["image_link", "image", "зображення"], "generatable": False, "label": "Image"},
    "product_type": {"aliases": ["product_type", "тип_товару"], "generatable": True, "label": "Product Type"},
    "google_product_category": {"aliases": ["google_product_category", "google_category", "gpc"], "generatable": True, "label": "Google Category"},
    "brand": {"aliases": ["brand", "manufacturer", "бренд", "виробник"], "generatable": True, "label": "Brand"},
    "gtin": {"aliases": ["gtin", "ean", "upc", "barcode"], "generatable": False, "label": "GTIN"},
    "mpn": {"aliases": ["mpn", "артикул"], "generatable": False, "label": "MPN"},
    "color": {"aliases": ["color", "colour", "колір"], "generatable": True, "label": "Color"},
    "material": {"aliases": ["material", "матеріал"], "generatable": True, "label": "Material"},
    "gender": {"aliases": ["gender", "стать"], "generatable": True, "label": "Gender"},
    "age_group": {"aliases": ["age_group", "вікова_група"], "generatable": True, "label": "Age Group"},
    "size": {"aliases": ["size", "розмір"], "generatable": True, "label": "Size"},
    "pattern": {"aliases": ["pattern", "візерунок", "малюнок"], "generatable": True, "label": "Pattern"},
    "condition": {"aliases": ["condition", "стан"], "generatable": True, "label": "Condition"},
    "price": {"aliases": ["price", "ціна"], "generatable": False, "label": "Price"},
    "availability": {"aliases": ["availability", "наявність"], "generatable": False, "label": "Availability"},
    "product_highlight": {"aliases": ["product_highlight", "highlight"], "generatable": True, "label": "Highlight"},
}

CORE_ATTRIBUTES = ["id", "title", "description", "product_type", "brand"]

STRUCTURABLE_ATTRIBUTES = [
    "brand", "gender", "age_group", "product_type", "color",
    "material", "pattern", "size", "google_product_category",
]

GENERATABLE_ATTRIBUTES = [k for k, v in GMC_ATTRIBUTES.items() if v["generatable"]]

DEFAULT_TITLE_ORDER = ["brand", "gender", "product_type", "color", "material", "pattern"]

ALL_ALIASES = {a for meta in GMC_ATTRIBUTES.values() for a in meta["aliases"]} | set(GMC_ATTRIBUTES.keys())


def detect_columns(headers):
    mapping = {}
    headers_lower = [str(h).lower().strip() if h else "" for h in headers]
    for key, meta in GMC_ATTRIBUTES.items():
        for alias in [key] + meta["aliases"]:
            if alias in headers_lower:
                mapping[key] = headers_lower.index(alias)
                break
    return mapping


def detect_header_row(first_row):
    if not first_row:
        return False
    cells = [str(c).lower().strip() for c in first_row if c is not None and str(c).strip()]
    if not cells:
        return False
    matches = sum(1 for c in cells if c in ALL_ALIASES)
    if matches >= 2:
        return True
    return cells[0] in ("id", "item_id", "offer_id", "товар_id")


def gmc_key_for_header(header):
    hl = str(header).lower().strip()
    for key, meta in GMC_ATTRIBUTES.items():
        if hl == key or hl in meta["aliases"]:
            return key
    return None


# =============================================================================
# MULTI-FORMAT FEED PARSING
# =============================================================================

def parse_feed_file(file_path, ext):
    if ext == "xlsx":
        return _parse_xlsx(file_path)
    elif ext == "csv":
        return _parse_csv(file_path)
    elif ext == "xml":
        return _parse_xml(file_path)
    raise ValueError(f"Непідтримуваний формат: .{ext}")


def _parse_xlsx(file_path):
    wb = openpyxl.load_workbook(file_path, read_only=True)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows:
        return [], []
    if detect_header_row(rows[0]):
        headers = [str(h) if h else "" for h in rows[0]]
        return headers, [list(r) for r in rows[1:]]
    headers = [f"col_{i}" for i in range(len(rows[0]))]
    return headers, [list(r) for r in rows]


def _parse_csv(file_path):
    for encoding in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            with open(file_path, "r", encoding=encoding, newline="") as f:
                sample = f.read(4096)
                f.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
                    delimiter = dialect.delimiter
                except csv.Error:
                    delimiter = ","
                reader = csv.reader(f, delimiter=delimiter)
                rows = [row for row in reader if any(cell.strip() for cell in row)]
            if not rows:
                return [], []
            if detect_header_row(rows[0]):
                headers = [h.strip() for h in rows[0]]
                return headers, rows[1:]
            headers = [f"col_{i}" for i in range(len(rows[0]))]
            return headers, rows
        except (UnicodeDecodeError, UnicodeError):
            continue
    raise ValueError("Не вдалося прочитати CSV (проблема з кодуванням)")


def _parse_xml(file_path):
    try:
        tree = DET.parse(file_path)
    except Exception as e:
        raise ValueError(f"Невалідний XML: {str(e)[:120]}")
    root = tree.getroot()

    items = root.findall(".//item")
    if not items:
        items = root.findall(".//{http://base.google.com/ns/1.0}item") or root.findall("item")
    if not items:
        raise ValueError("XML не містить <item> елементів (очікується формат Google Shopping)")

    all_keys = []
    seen_keys = set()
    parsed_items = []
    for item in items:
        item_data = {}
        for child in item:
            tag = child.tag
            if "}" in tag:
                tag = tag.split("}")[-1]
            tag = tag.lower()
            val = (child.text or "").strip()
            item_data[tag] = val
            if tag not in seen_keys:
                seen_keys.add(tag)
                all_keys.append(tag)
        parsed_items.append(item_data)

    headers = all_keys
    data_rows = [[item.get(k, "") for k in headers] for item in parsed_items]
    return headers, data_rows


# =============================================================================
# OUTPUT WRITING
# =============================================================================

def _xml_tag(name):
    tag = re.sub(r"[^\w\-.]", "_", str(name).strip())
    if not tag or tag[0].isdigit() or tag.startswith(("-", ".")):
        tag = "f_" + tag
    return tag


def write_output(output_rows, out_headers, output_format, output_path):
    if output_format == "csv":
        with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(out_headers)
            for row in output_rows:
                writer.writerow(["" if v is None else v for v in row])

    elif output_format == "xml":
        ET.register_namespace("g", "http://base.google.com/ns/1.0")
        rss = ET.Element("rss", {"version": "2.0"})
        channel = ET.SubElement(rss, "channel")
        ET.SubElement(channel, "title").text = "FeedGen Generated Feed"
        ET.SubElement(channel, "link").text = "https://merchants.google.com"
        ET.SubElement(channel, "description").text = f"Generated by FeedGen, {datetime.now().date().isoformat()}"

        header_tags = []
        for h in out_headers:
            gmc_key = gmc_key_for_header(h)
            if gmc_key:
                header_tags.append(f"{{http://base.google.com/ns/1.0}}{gmc_key}")
            else:
                header_tags.append(_xml_tag(h))

        for row in output_rows:
            item = ET.SubElement(channel, "item")
            for tag, val in zip(header_tags, row):
                el = ET.SubElement(item, tag)
                el.text = "" if val is None else str(val)
        tree = ET.ElementTree(rss)
        tree.write(output_path, encoding="utf-8", xml_declaration=True)

    else:  # xlsx
        out_wb = Workbook()
        out_ws = out_wb.active
        out_ws.title = "Generated Feed"
        hfont = Font(bold=True, size=11, name="Arial")
        hfill = PatternFill("solid", start_color="D9E1F2")
        for ci, h in enumerate(out_headers, 1):
            c = out_ws.cell(row=1, column=ci, value=h)
            c.font, c.fill = hfont, hfill
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for ri, rd in enumerate(output_rows, 2):
            for ci, val in enumerate(rd, 1):
                if val is not None and not isinstance(val, (str, int, float, bool)):
                    val = str(val)
                c = out_ws.cell(row=ri, column=ci, value=val)
                c.font = Font(name="Arial", size=10)
                c.alignment = Alignment(vertical="top", wrap_text=True)
        out_wb.save(str(output_path))


# =============================================================================
# ROUTES: pages, auth, schema, analyze
# =============================================================================

@app.get("/")
async def index(request: Request):
    return templates.TemplateResponse(request=request, name="index.html")


@app.get("/login")
async def login_page(request: Request):
    if not AUTH_ENABLED:
        return RedirectResponse(url="/")
    return templates.TemplateResponse(request=request, name="login.html")


@app.post("/api/login")
async def api_login(request: Request, token: str = Form(...)):
    ip = request.client.host if request.client else "unknown"
    now = time.time()
    attempts = [t for t in login_attempts.get(ip, []) if now - t < LOGIN_WINDOW]
    if len(attempts) >= LOGIN_MAX_ATTEMPTS:
        logger.warning("login rate limit hit for ip=%s", ip)
        raise HTTPException(429, "Забагато спроб входу. Спробуй за 5 хвилин.")

    label = verify_token(token.strip())
    if not label:
        attempts.append(now)
        login_attempts[ip] = attempts
        return RedirectResponse(url="/login?error=1", status_code=303)

    login_attempts.pop(ip, None)
    logger.info("login ok: client=%s", label)
    response = RedirectResponse(url="/", status_code=303)
    response.set_cookie(
        key="feedgen_token",
        value=token.strip(),
        httponly=True,
        secure=COOKIE_SECURE,
        samesite="lax",
        max_age=60 * 60 * 24 * 30,
    )
    return response


@app.post("/api/logout")
async def api_logout():
    response = RedirectResponse(url="/login", status_code=303)
    response.delete_cookie("feedgen_token")
    return response


@app.get("/api/me")
async def api_me(request: Request):
    if not AUTH_ENABLED:
        return {"auth_enabled": False, "client": "open"}
    label = get_client(request)
    return {"auth_enabled": True, "client": label or "unknown"}


@app.get("/api/schema")
async def get_schema():
    return {
        "structurable": [
            {"key": k, "label": GMC_ATTRIBUTES[k]["label"]}
            for k in STRUCTURABLE_ATTRIBUTES
        ],
        "generatable": [
            {"key": k, "label": GMC_ATTRIBUTES[k]["label"]}
            for k in GENERATABLE_ATTRIBUTES
        ],
        "default_title_order": DEFAULT_TITLE_ORDER,
        "limits": {
            "max_feed_size_mb": MAX_FEED_SIZE // (1024 * 1024),
            "max_unique_products": MAX_UNIQUE_PRODUCTS,
            "preview_size": PREVIEW_DEFAULT_SIZE,
        },
    }


async def _read_upload_limited(file: UploadFile, max_size: int, what: str) -> bytes:
    chunks = []
    size = 0
    while True:
        chunk = await file.read(1024 * 1024)
        if not chunk:
            break
        size += len(chunk)
        if size > max_size:
            raise HTTPException(413, f"{what} завеликий (ліміт {max_size // (1024 * 1024)} МБ)")
        chunks.append(chunk)
    return b"".join(chunks)


@app.post("/api/analyze")
async def analyze_feed(request: Request, file: UploadFile = File(...)):
    await cleanup_old_files()

    original_name = file.filename or "feed"
    ext = original_name.lower().rsplit(".", 1)[-1] if "." in original_name else ""
    if ext not in ALLOWED_INPUT_EXTS:
        raise HTTPException(400, "Підтримуються формати: XLSX, CSV, XML")

    content = await _read_upload_limited(file, MAX_FEED_SIZE, "Файл")
    if not content:
        raise HTTPException(400, "Файл порожній")

    file_id = uuid.uuid4().hex
    tmp_path = UPLOAD_DIR / f"{file_id}.{ext}"
    tmp_path.write_bytes(content)

    client = get_client(request) or "open"

    try:
        headers, data_rows = await asyncio.to_thread(parse_feed_file, tmp_path, ext)
        if not data_rows:
            raise HTTPException(400, "Файл порожній або не містить даних")

        col_map = detect_columns(headers)
        title_idx = col_map.get("title", 0)
        titles = Counter(
            row[title_idx] for row in data_rows
            if len(row) > title_idx and row[title_idx]
        )

        missing_generatable = [k for k in GENERATABLE_ATTRIBUTES if k not in col_map]
        present_attributes = list(col_map.keys())

        sample = {}
        for key, idx in col_map.items():
            if key in ("id", "description", "link", "image_link"):
                continue
            vals = set()
            for row in data_rows[:100]:
                if len(row) > idx and row[idx]:
                    vals.add(str(row[idx]))
            if vals:
                sample[key] = sorted(vals)[:10]

        uploaded_files[file_id] = {
            "path": str(tmp_path),
            "ext": ext,
            "client": client,
            "filename": original_name,
            "created": datetime.now().isoformat(),
        }
        logger.info("analyze ok: client=%s file=%s rows=%d unique=%d",
                    client, original_name, len(data_rows), len(titles))

        return {
            "file_id": file_id,
            "filename": original_name,
            "format": ext,
            "total_rows": len(data_rows),
            "unique_titles": len(titles),
            "columns": headers[:40],
            "detected_columns": col_map,
            "present_attributes": present_attributes,
            "missing_generatable": missing_generatable,
            "sample_values": sample,
            "over_limit": len(titles) > MAX_UNIQUE_PRODUCTS,
            "max_unique_products": MAX_UNIQUE_PRODUCTS,
        }
    except HTTPException:
        tmp_path.unlink(missing_ok=True)
        raise
    except Exception as e:
        tmp_path.unlink(missing_ok=True)
        logger.warning("analyze failed: client=%s file=%s err=%s", client, original_name, e)
        raise HTTPException(400, f"Помилка читання файлу: {str(e)[:200]}")


# =============================================================================
# SHARED VALIDATION for preview & generate
# =============================================================================

async def _validate_generation_params(request, file_id, config, config_text,
                                      model, language, column_map, title_order,
                                      generate_attributes):
    """Validates all form params shared by /api/preview and /api/generate.
    Returns a dict of validated values. Raises HTTPException on any problem."""
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        raise HTTPException(500, "ANTHROPIC_API_KEY не налаштований на сервері")

    client_label = get_client(request) or "open"

    file_id = (file_id or "").strip()
    if not FILE_ID_RE.match(file_id):
        raise HTTPException(400, "Невалідний file_id")

    meta = uploaded_files.get(file_id)
    if not meta or not Path(meta["path"]).exists():
        raise HTTPException(404, "Файл не знайдено (можливо, сервер перезапускався). Завантаж фід ще раз.")
    if AUTH_ENABLED and meta["client"] != client_label:
        raise HTTPException(404, "Файл не знайдено. Завантаж фід ще раз.")

    if model not in ALLOWED_MODELS:
        raise HTTPException(400, f"Невалідна модель. Доступні: {', '.join(sorted(ALLOWED_MODELS))}")
    if language not in ALLOWED_LANGUAGES:
        raise HTTPException(400, "Невалідна мова (uk або en)")

    def parse_json_form(value, name, expected_type):
        if not value:
            return None
        try:
            parsed = json.loads(value)
        except json.JSONDecodeError:
            raise HTTPException(400, f"Невалідний JSON у полі {name}")
        if not isinstance(parsed, expected_type):
            raise HTTPException(400, f"Поле {name} має неправильний тип")
        return parsed

    niche_config = None
    if config:
        raw = await _read_upload_limited(config, MAX_CONFIG_SIZE, "Конфіг")
        try:
            niche_config = json.loads(raw)
        except (json.JSONDecodeError, UnicodeDecodeError):
            raise HTTPException(400, "Конфіг ніші - невалідний JSON")
    elif config_text:
        if len(config_text) > MAX_CONFIG_SIZE:
            raise HTTPException(413, "Конфіг завеликий")
        try:
            niche_config = json.loads(config_text)
        except json.JSONDecodeError:
            raise HTTPException(400, "Конфіг ніші - невалідний JSON")
    if niche_config is not None and not isinstance(niche_config, dict):
        raise HTTPException(400, "Конфіг ніші має бути JSON-об'єктом")

    col_map = parse_json_form(column_map, "column_map", dict)
    if col_map is not None:
        col_map = {
            k: v for k, v in col_map.items()
            if k in GMC_ATTRIBUTES and isinstance(v, int) and v >= 0
        }

    order = parse_json_form(title_order, "title_order", list) or DEFAULT_TITLE_ORDER
    order = [a for a in order if a in STRUCTURABLE_ATTRIBUTES] or DEFAULT_TITLE_ORDER

    gen_attrs = parse_json_form(generate_attributes, "generate_attributes", list) or []
    gen_attrs = [a for a in gen_attrs if a in GENERATABLE_ATTRIBUTES]

    return {
        "api_key": api_key,
        "client": client_label,
        "meta": meta,
        "niche_config": niche_config,
        "col_map": col_map,
        "title_order": order,
        "gen_attrs": gen_attrs,
    }


def extract_banned_words(niche_config):
    """Collects banned words from all supported config locations."""
    if not niche_config:
        return []
    cands = list(niche_config.get("banned_words", []) or [])
    tov = niche_config.get("tone_of_voice", {})
    if isinstance(tov, dict):
        cands += tov.get("banned_words", []) or []
        cands += tov.get("word_taboos", []) or []
    bc = niche_config.get("brand_config", {})
    if isinstance(bc, dict):
        cands += bc.get("banned_words", []) or []
    return [w for w in cands if isinstance(w, str) and len(w) > 2]


def build_unique_products(data_rows, col_map, gen_attrs):
    """Deduplicates rows. Returns (unique_products dict, row_keys list)."""
    include_attrs = bool(gen_attrs)
    unique_products = {}
    row_keys = []
    for row in data_rows:
        key = dedup_key(row, col_map, include_attrs)
        if key not in unique_products:
            unique_products[key] = extract_product_data(row, col_map)
        row_keys.append(key)
    return unique_products, row_keys


def make_system_blocks(system_text, model):
    if model in CACHE_SUPPORTED_MODELS:
        return [{"type": "text", "text": system_text, "cache_control": {"type": "ephemeral"}}], True
    return system_text, False


# =============================================================================
# ROUTES: preview (new in v4)
# =============================================================================

@app.post("/api/preview")
async def preview_generation(
    request: Request,
    file_id: str = Form(...),
    config: UploadFile = File(None),
    config_text: str = Form(None),
    model: str = Form("claude-sonnet-5"),
    language: str = Form("uk"),
    column_map: str = Form(None),
    title_order: str = Form(None),
    generate_attributes: str = Form(None),
    sample_size: int = Form(PREVIEW_DEFAULT_SIZE),
):
    """Generates a small sample synchronously so the user can check quality
    and see a cost estimate BEFORE committing to the full (paid) run."""
    import anthropic

    p = await _validate_generation_params(
        request, file_id, config, config_text, model, language,
        column_map, title_order, generate_attributes,
    )
    sample_size = max(1, min(int(sample_size or PREVIEW_DEFAULT_SIZE), PREVIEW_MAX_SIZE))

    meta = p["meta"]
    headers, data_rows = await asyncio.to_thread(parse_feed_file, Path(meta["path"]), meta["ext"])
    col_map = p["col_map"] or detect_columns(headers)
    unique_products, _ = build_unique_products(data_rows, col_map, p["gen_attrs"])
    total_unique = len(unique_products)
    if total_unique == 0:
        raise HTTPException(400, "У фіді не знайдено товарів для генерації")

    system_text = build_system_prompt(p["niche_config"], language, p["title_order"], p["gen_attrs"])
    system_blocks, use_caching = make_system_blocks(system_text, model)
    banned_words = extract_banned_words(p["niche_config"])

    client = anthropic.AsyncAnthropic(api_key=p["api_key"])
    sample_items = list(unique_products.values())[:sample_size]

    results = []
    usages = []
    seen_titles = set()

    # First call warms the prompt cache, the rest run in parallel (max 3)
    sem = asyncio.Semaphore(3)

    async def gen_one(product):
        async with sem:
            return product, await call_claude_with_retry(
                client, model, system_blocks, product, p["gen_attrs"]
            )

    ordered = []
    if sample_items:
        first = sample_items[0]
        ordered.append((first, await call_claude_with_retry(
            client, model, system_blocks, first, p["gen_attrs"])))
        tasks = [asyncio.create_task(gen_one(pr)) for pr in sample_items[1:]]
        for t in tasks:
            ordered.append(await t)

    for product, (gt, gd, gattrs, error, usage) in ordered:
        if usage:
            usages.append(usage)
        if gt is None:
            results.append({
                "id": str(product["id"]),
                "original_title": product["title"],
                "generated_title": None,
                "generated_description": None,
                "attributes": {},
                "score": 1,
                "comment": f"помилка: {error}",
            })
            continue
        score, comment = validate_result(gt, gd, banned_words, seen_titles)
        seen_titles.add(gt)
        results.append({
            "id": str(product["id"]),
            "original_title": product["title"],
            "generated_title": gt,
            "generated_description": gd,
            "attributes": gattrs,
            "score": score,
            "comment": comment,
        })

    # Cost estimate from REAL preview usage, scaled to the full feed.
    # No-cache upper bound; prompt caching typically cuts input cost a lot.
    estimate = None
    if usages and model in MODEL_PRICING:
        avg_in = sum(u["input_tokens"] for u in usages) / len(usages)
        avg_out = sum(u["output_tokens"] for u in usages) / len(usages)
        in_price, out_price = MODEL_PRICING[model]
        est = total_unique * (avg_in * in_price + avg_out * out_price) / 1_000_000
        estimate = {
            "total_unique": total_unique,
            "avg_input_tokens": round(avg_in),
            "avg_output_tokens": round(avg_out),
            "est_cost_usd_max": round(est, 2),
            "note": "Верхня межа без урахування prompt caching - реальна вартість зазвичай суттєво нижча",
        }

    logger.info("preview ok: client=%s file=%s sample=%d unique=%d",
                p["client"], meta["filename"], len(results), total_unique)

    return {
        "results": results,
        "total_unique": total_unique,
        "sample_size": len(results),
        "estimate": estimate,
        "over_limit": total_unique > MAX_UNIQUE_PRODUCTS,
        "max_unique_products": MAX_UNIQUE_PRODUCTS,
    }


# =============================================================================
# ROUTES: generate / status / cancel / download
# =============================================================================

@app.post("/api/generate")
async def start_generation(
    request: Request,
    background_tasks: BackgroundTasks,
    file_id: str = Form(...),
    config: UploadFile = File(None),
    config_text: str = Form(None),
    model: str = Form("claude-sonnet-5"),
    language: str = Form("uk"),
    column_map: str = Form(None),
    title_order: str = Form(None),
    generate_attributes: str = Form(None),
    output_format: str = Form("xlsx"),
):
    p = await _validate_generation_params(
        request, file_id, config, config_text, model, language,
        column_map, title_order, generate_attributes,
    )
    if output_format not in ALLOWED_OUTPUT_FORMATS:
        raise HTTPException(400, "Невалідний формат експорту (xlsx, csv або xml)")

    job_id = uuid.uuid4().hex[:12]
    jobs[job_id] = {
        "status": "processing", "progress": 0, "total": 0,
        "message": "Підготовка...", "created": datetime.now().isoformat(),
        "client": p["client"], "cancel_requested": False,
        "output_file": None, "errors": [], "stats": {},
    }
    logger.info("generate start: job=%s client=%s model=%s format=%s gen_attrs=%s",
                job_id, p["client"], model, output_format, p["gen_attrs"])

    background_tasks.add_task(
        run_generation, job_id, p["meta"]["path"], p["meta"]["ext"], p["api_key"],
        model, language, p["niche_config"], p["col_map"], p["title_order"],
        p["gen_attrs"], output_format,
    )
    return {"job_id": job_id}


def _get_own_job(job_id: str, request: Request):
    if not JOB_ID_RE.match(job_id or ""):
        raise HTTPException(404, "Завдання не знайдено")
    job = jobs.get(job_id)
    if not job:
        raise HTTPException(404, "Завдання не знайдено")
    if AUTH_ENABLED:
        client = get_client(request)
        if job.get("client") != client:
            raise HTTPException(404, "Завдання не знайдено")
    return job


@app.get("/api/status/{job_id}")
async def get_status(job_id: str, request: Request):
    job = _get_own_job(job_id, request)
    return {k: v for k, v in job.items() if k != "output_file"}


@app.post("/api/cancel/{job_id}")
async def cancel_job(job_id: str, request: Request):
    """Requests cancellation of a running job. Already-generated products
    are kept and written to a partial output file."""
    job = _get_own_job(job_id, request)
    if job["status"] != "processing":
        raise HTTPException(400, "Завдання вже завершене - скасовувати нічого")
    job["cancel_requested"] = True
    job["message"] = "Скасовую... зберігаю вже згенероване"
    logger.info("cancel requested: job=%s", job_id)
    return {"ok": True}


@app.get("/api/download/{job_id}")
async def download_result(job_id: str, request: Request):
    job = _get_own_job(job_id, request)
    if job["status"] not in ("done", "cancelled") or not job["output_file"]:
        raise HTTPException(400, "Файл ще не готовий")
    fmt = job.get("stats", {}).get("output_format", "xlsx")
    media_types = {
        "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "csv": "text/csv",
        "xml": "application/xml",
    }
    suffix = "_partial" if job["status"] == "cancelled" else ""
    return FileResponse(
        job["output_file"],
        filename=f"generated_feed_{job_id}{suffix}.{fmt}",
        media_type=media_types.get(fmt, "application/octet-stream"),
    )


@app.get("/api/health")
async def health():
    return {"status": "ok", "version": "4.0"}


# =============================================================================
# GENERATION ENGINE
# =============================================================================

def build_system_prompt(niche_config, language, title_order, gen_attrs):
    if niche_config:
        base = json.dumps(niche_config, ensure_ascii=False, indent=2)
        extra = "\n\n--- STRUCTURE & GENERATION SETTINGS ---\n"
        extra += f"Title attribute order: {' -> '.join(title_order)}\n"
        if gen_attrs:
            extra += f"Generate these missing attributes if possible: {', '.join(gen_attrs)}\n"
        return base + extra

    lang_name = "українська" if language == "uk" else "English"
    order_str = " -> ".join(title_order)

    prompt = f"""You are a specialized AI copywriter for Google Merchant Center product feeds.
Generation language: {lang_name}.

TASK — for each product, based ONLY on the provided attributes, generate:
1. An optimized title (max 150 chars). Follow this attribute order: {order_str}.
   Most important info in the first 70 chars. Skip attributes that are absent.
2. An optimized description (50-80 words) — factual, structured, highlighting key attributes.
"""

    if gen_attrs:
        attrs_list = ", ".join(gen_attrs)
        prompt += f"""3. Infer and fill these MISSING attributes when they can be reliably determined
   from the title/description/other data: {attrs_list}.
   Only fill an attribute if you are confident. Leave empty ("") if uncertain.
   Never invent values that contradict the data.
"""

    prompt += """
RULES:
- Use ONLY provided data. Never invent specifications.
- Each title must be unique.
- Maintain grammatical correctness in the target language (agreement, cases).
- Adapt tone to the product category implied by the attributes.

RESPONSE FORMAT — respond ONLY with valid JSON:
"""
    if gen_attrs:
        attrs_json = ", ".join(f'"{a}": "..."' for a in gen_attrs)
        prompt += f'{{"title": "...", "description": "...", "attributes": {{{attrs_json}}}}}'
    else:
        prompt += '{"title": "...", "description": "..."}'

    return prompt


def extract_product_data(row, col_map):
    def get(key, default=""):
        idx = col_map.get(key)
        if idx is not None and len(row) > idx and row[idx]:
            return row[idx]
        return default

    data = {
        "id": get("id"),
        "title": str(get("title")),
        "description": str(get("description"))[:500],
        "product_type": get("product_type"),
        "brand": get("brand"),
        "attributes": {},
    }
    for key in GMC_ATTRIBUTES:
        if key in CORE_ATTRIBUTES:
            continue
        val = get(key)
        if val:
            data["attributes"][key] = str(val)
    return data


def dedup_key(row, col_map, include_attrs):
    """Base: (title, description). With attribute generation, existing attribute
    values are included too - otherwise product variants would merge and one
    variant's generated attributes would spread to all of them."""
    def get(key):
        idx = col_map.get(key)
        if idx is not None and len(row) > idx and row[idx]:
            return str(row[idx])
        return ""

    base = (get("title"), get("description"))
    if not include_attrs:
        return base
    attr_part = tuple(
        (k, get(k)) for k in sorted(GMC_ATTRIBUTES)
        if k not in ("id", "title", "description", "link", "image_link") and get(k)
    )
    return base + (attr_part,)


def parse_claude_response(text, gen_attrs=None):
    text = text.strip()
    candidates = []

    try:
        candidates.append(json.loads(text))
    except json.JSONDecodeError:
        pass
    cb = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", text, re.DOTALL)
    if cb:
        try:
            candidates.append(json.loads(cb.group(1)))
        except json.JSONDecodeError:
            pass
    f, l = text.find("{"), text.rfind("}")
    if f != -1 and l > f:
        try:
            candidates.append(json.loads(text[f:l + 1]))
        except json.JSONDecodeError:
            pass

    for result in candidates:
        if isinstance(result, dict) and "title" in result:
            attrs = result.get("attributes", {}) if gen_attrs else {}
            if not isinstance(attrs, dict):
                attrs = {}
            return str(result["title"]), str(result.get("description", "")), attrs
    return None, None, {}


async def call_claude_with_retry(client, model, system_blocks, product_data, gen_attrs, max_retries=3):
    """Calls Claude API (async) with prompt caching and exponential backoff.
    Returns (title, description, attributes, error, usage). usage is
    {"input_tokens": int, "output_tokens": int} or None on failure."""
    import anthropic

    attr_lines = [
        f"Current title: {product_data['title']}",
        f"Description: {product_data['description'][:300]}",
    ]
    if product_data.get("product_type"):
        attr_lines.append(f"Product type: {product_data['product_type']}")
    if product_data.get("brand"):
        attr_lines.append(f"Brand: {product_data['brand']}")
    for k, v in product_data.get("attributes", {}).items():
        attr_lines.append(f"{k.replace('_', ' ').capitalize()}: {v}")

    user_message = (
        "Generate optimized content for this product.\n\n"
        + "\n".join(attr_lines)
    )

    # Claude Sonnet 5 rejects non-default sampling params (temperature/top_p/top_k)
    # with a 400 error - this is a real API behavior change, not a bug on our
    # side. Older models still accept and benefit from a lower temperature for
    # more consistent, less "creative" product copy, so we keep it for them.
    create_kwargs = dict(
        model=model,
        max_tokens=1500,
        system=system_blocks,
        messages=[{"role": "user", "content": user_message}],
    )
    if model not in SONNET5_NO_SAMPLING_PARAMS:
        create_kwargs["temperature"] = 0.3

    for attempt in range(max_retries):
        try:
            response = await client.messages.create(**create_kwargs)
            usage = None
            try:
                u = response.usage
                cache_read = getattr(u, "cache_read_input_tokens", 0) or 0
                cache_write = getattr(u, "cache_creation_input_tokens", 0) or 0
                usage = {
                    "input_tokens": (u.input_tokens or 0) + cache_read + cache_write,
                    "output_tokens": u.output_tokens or 0,
                }
            except Exception:
                pass
            text = response.content[0].text.strip()
            title, desc, attrs = parse_claude_response(text, gen_attrs)
            if title is not None:
                return title, desc, attrs, None, usage
            return None, None, {}, f"parse_failed: {text[:80]}", usage

        except anthropic.RateLimitError:
            await asyncio.sleep((2 ** attempt) * 2)
            continue
        except anthropic.APIStatusError as e:
            if e.status_code in (429, 500, 502, 503, 529):
                await asyncio.sleep((2 ** attempt) * 2)
                continue
            return None, None, {}, f"api_error: {str(e)[:120]}", None
        except Exception as e:
            return None, None, {}, f"error: {str(e)[:120]}", None

    return None, None, {}, "max_retries_exceeded", None


def validate_result(title, description, banned_words, seen_titles):
    score = 5
    comments = []
    if len(title) > 150:
        score = min(score, 3)
        comments.append(f"title {len(title)} chars (>150)")
    tl = title.lower()
    dl = (description or "").lower()
    for w in banned_words:
        wl = w.lower()
        if wl in tl:
            score = min(score, 2)
            comments.append(f"banned word '{w}' in title")
        elif wl in dl:
            score = min(score, 3)
            comments.append(f"banned word '{w}' in description")
    if title in seen_titles:
        score = min(score, 3)
        comments.append("duplicate title")
    return score, "; ".join(comments)


async def run_generation(job_id, input_path, ext, api_key, model, language,
                         niche_config, col_map_override, title_order, gen_attrs, output_format):
    """Background generation task. Async client + bounded parallelism.
    Supports cancellation: partial results are written and downloadable."""
    import anthropic

    try:
        client = anthropic.AsyncAnthropic(api_key=api_key)
        system_text = build_system_prompt(niche_config, language, title_order, gen_attrs)
        system_blocks, use_caching = make_system_blocks(system_text, model)
        banned_words = extract_banned_words(niche_config)

        headers, data_rows = await asyncio.to_thread(parse_feed_file, Path(input_path), ext)
        col_map = col_map_override or detect_columns(headers)

        unique_products, row_keys = build_unique_products(data_rows, col_map, gen_attrs)
        total_unique = len(unique_products)
        if total_unique > MAX_UNIQUE_PRODUCTS:
            jobs[job_id]["status"] = "error"
            jobs[job_id]["message"] = (
                f"Забагато унікальних товарів: {total_unique} (ліміт {MAX_UNIQUE_PRODUCTS}). "
                f"Розбий фід на частини або підніми MAX_UNIQUE_PRODUCTS."
            )
            return

        jobs[job_id]["total"] = total_unique
        jobs[job_id]["message"] = f"Генерація {total_unique} унікальних товарів..."

        generated = {}
        seen_titles = set()
        fallback_count = 0
        sem = asyncio.Semaphore(GENERATION_CONCURRENCY)

        async def gen_one(key, product):
            async with sem:
                result = await call_claude_with_retry(
                    client, model, system_blocks, product, gen_attrs
                )
                return key, product, result

        items = list(unique_products.items())

        def handle_result(key, product, result):
            nonlocal fallback_count
            gt, gd, gattrs, error, _usage = result
            if gt is None:
                gt = product["title"]
                gd = str(product.get("description", ""))
                gattrs = {}
                fallback_count += 1
                if len(jobs[job_id]["errors"]) < 50:
                    jobs[job_id]["errors"].append(f"ID {product['id']}: {error}")
                score, comment = 1, f"fallback: {error}"
            else:
                score, comment = validate_result(gt, gd, banned_words, seen_titles)
            seen_titles.add(gt)
            generated[key] = (gt, gd, gattrs, score, comment)
            jobs[job_id]["progress"] = len(generated)
            jobs[job_id]["message"] = f"Оброблено {len(generated)}/{total_unique}"

        cancelled = False

        # 1) cache warm-up: one request writes the cache before parallel reads
        if items and use_caching:
            key0, product0 = items[0]
            result0 = await call_claude_with_retry(client, model, system_blocks, product0, gen_attrs)
            handle_result(key0, product0, result0)
            items = items[1:]
            if jobs[job_id].get("cancel_requested"):
                cancelled = True

        # 2) parallel generation with cancellation checks between completions
        if not cancelled:
            tasks = [asyncio.create_task(gen_one(k, p)) for k, p in items]
            try:
                for fut in asyncio.as_completed(tasks):
                    key, product, result = await fut
                    handle_result(key, product, result)
                    if jobs[job_id].get("cancel_requested"):
                        cancelled = True
                        break
            finally:
                if cancelled:
                    for t in tasks:
                        t.cancel()
                    await asyncio.gather(*tasks, return_exceptions=True)

        # Output: ALL original columns preserved + generated columns appended.
        # On cancellation, unprocessed rows keep empty generated fields.
        gen_headers = ["generated_title", "generated_description"]
        gen_headers += [f"generated_{a}" for a in gen_attrs]
        gen_headers += ["quality_score", "quality_comment"]
        out_headers = [str(h) for h in headers] + gen_headers

        n_cols = len(headers)
        n_gen_attrs = len(gen_attrs)
        output_rows = []
        skipped_rows = 0
        for i, row in enumerate(data_rows):
            base = list(row)[:n_cols] + [""] * max(0, n_cols - len(row))
            g = generated.get(row_keys[i])
            if g is not None:
                out_row = base + [g[0], g[1]] + [g[2].get(a, "") for a in gen_attrs] + [g[3], g[4]]
            else:
                skipped_rows += 1
                out_row = base + ["", ""] + [""] * n_gen_attrs + ["", "не оброблено (скасовано)"]
            output_rows.append(out_row)

        output_path = UPLOAD_DIR / f"generated_{job_id}.{output_format}"
        await asyncio.to_thread(write_output, output_rows, out_headers, output_format, str(output_path))

        score_dist = Counter(g[3] for g in generated.values())
        unique_gen = len(set(g[0] for g in generated.values()))

        jobs[job_id]["status"] = "cancelled" if cancelled else "done"
        jobs[job_id]["output_file"] = str(output_path)
        jobs[job_id]["stats"] = {
            "total_rows": len(output_rows),
            "unique_products": total_unique,
            "processed_unique": len(generated),
            "skipped_rows": skipped_rows,
            "unique_generated": unique_gen,
            "fallback_count": fallback_count,
            "caching_enabled": use_caching,
            "concurrency": GENERATION_CONCURRENCY,
            "generated_attributes": gen_attrs,
            "output_format": output_format,
            "score_distribution": {str(k): v for k, v in score_dist.items()},
        }
        if cancelled:
            jobs[job_id]["message"] = (
                f"Скасовано. Збережено {len(generated)}/{total_unique} унікальних товарів "
                f"({skipped_rows} рядків без генерації). Частковий файл готовий."
            )
            logger.info("generate cancelled: job=%s processed=%d/%d",
                        job_id, len(generated), total_unique)
        else:
            jobs[job_id]["message"] = (
                f"Готово! {len(output_rows)} товарів. Fallback: {fallback_count}. "
                f"Формат: {output_format.upper()}"
            )
            logger.info("generate done: job=%s rows=%d unique=%d fallback=%d",
                        job_id, len(output_rows), total_unique, fallback_count)

    except Exception as e:
        logger.error("generate failed: job=%s\n%s", job_id, traceback.format_exc())
        jobs[job_id]["status"] = "error"
        jobs[job_id]["message"] = f"Помилка: {str(e)[:300]}"
